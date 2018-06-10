import itertools
import openpyxl
import csv


def main():   
    EXCEL_FILE = 'hrvatski_dinar.xlsx'
    wb = openpyxl.load_workbook(EXCEL_FILE)

    with open("./hrvatski_dinar.csv", "w") as csv_file:
 
        fieldnames = ['date', 'middle_exchang_rate', 'currency_code']
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()

        rowStart = 6
        columnStart = 1
    
        for sheet in wb.worksheets:
            for offset in itertools.count():
                rowNum = rowStart + offset
                columnNum = columnStart
                
                date = sheet.cell(row=rowNum, column=columnNum ).value
                middle_exchang_rate = sheet.cell(row=rowNum, column=columnNum+1 ).value
                currency_code = sheet.cell(row=rowNum, column=columnNum+2 ).value

                if date == None:
                    break
                else:
                    writer.writerow({'date': date, 'middle_exchang_rate': middle_exchang_rate, 'currency_code': currency_code})


if __name__ == "__main__":
    main()

